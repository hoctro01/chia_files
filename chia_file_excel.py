#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tool Chia File Excel - Split Excel files into smaller chunks
============================================================
Chia nội dung file Excel (.xls) thành các file Excel (.xls)
có tối đa 8000 dòng dữ liệu, GIỮ NGUYÊN FORM MẪU GỐC.

Cách hoạt động:
    - Dùng xlutils.copy để sao chép NGUYÊN BẢN format từ file gốc
    - Giữ nguyên: font, size, bold, border, background color, merged cells,
      column widths, row heights, header rows
    - Tự động nhận diện header rows (các dòng trước dữ liệu)

Yêu cầu:
    pip install xlrd==1.2.0 xlwt==1.3.0 xlutils openpyxl

Sử dụng:
    python chia_file_excel.py                    # GUI mode
    python chia_file_excel.py "file.xls"         # Command line mode
"""

import os
import sys
import math
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

# ============================================================
# CONSTANTS
# ============================================================
MAX_ROWS_PER_FILE = 8000  # Số dòng dữ liệu tối đa mỗi file (không tính header)

# ============================================================
# CORE LOGIC - GIỮ NGUYÊN FORMAT GỐC
# ============================================================

def detect_header_rows(sheet):
    """
    Tự động nhận diện số dòng header (title + column headers).
    Header rows là các dòng đầu tiên TRƯỚC khi dữ liệu bắt đầu.
    Dữ liệu bắt đầu khi cột đầu tiên (STT) có giá trị số.
    """
    for row_idx in range(min(20, sheet.nrows)):  # Check first 20 rows max
        cell_value = sheet.cell_value(row_idx, 0)
        # Data starts when col 0 has a number (STT = 1, 2, 3...)
        if isinstance(cell_value, (int, float)) and cell_value >= 1:
            return row_idx
    # Fallback: assume row 0 is header
    return 1


def split_xls_preserve_format(filepath, max_rows=MAX_ROWS_PER_FILE, progress_callback=None):
    """
    Chia file .xls giữ nguyên format gốc bằng xlutils.copy.
    Mỗi file output sẽ có ĐÚNG format như file gốc.
    """
    import xlrd
    import xlwt
    from xlutils.copy import copy as xlutils_copy

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Không tìm thấy file: {filepath}")

    if progress_callback:
        progress_callback(0, 100, "Đang đọc file nguồn...")

    # === Đọc file gốc với formatting_info ===
    source_wb = xlrd.open_workbook(filepath, formatting_info=True)
    source_sheet = source_wb.sheet_by_index(0)

    total_rows = source_sheet.nrows
    total_cols = source_sheet.ncols

    if total_rows <= 1:
        raise ValueError("File Excel trống hoặc chỉ có header!")

    # Nhận diện header rows
    header_row_count = detect_header_rows(source_sheet)
    data_row_count = total_rows - header_row_count

    if progress_callback:
        progress_callback(5, 100,
            f"File có {total_rows:,} dòng tổng cộng: "
            f"{header_row_count} dòng header + {data_row_count:,} dòng dữ liệu")

    if data_row_count <= max_rows:
        if progress_callback:
            progress_callback(100, 100,
                f"File chỉ có {data_row_count:,} dòng dữ liệu, không cần chia!")
        return []

    # Tính số file cần tạo
    num_files = math.ceil(data_row_count / max_rows)

    # Tạo thư mục output
    source_dir = os.path.dirname(os.path.abspath(filepath))
    source_name = os.path.splitext(os.path.basename(filepath))[0]
    output_dir = os.path.join(source_dir, f"{source_name}_split")
    os.makedirs(output_dir, exist_ok=True)

    # === Thu thập thông tin format từ file gốc ===
    # Column widths
    col_widths = {}
    for col_idx in range(total_cols):
        ci = source_sheet.colinfo_map.get(col_idx)
        if ci:
            col_widths[col_idx] = ci.width

    # Row heights cho header
    header_row_heights = {}
    for row_idx in range(header_row_count):
        ri = source_sheet.rowinfo_map.get(row_idx)
        if ri:
            header_row_heights[row_idx] = ri.height

    # Merged cells trong header area
    header_merges = []
    for (rlo, rhi, clo, chi) in source_sheet.merged_cells:
        if rlo < header_row_count:
            header_merges.append((rlo, rhi, clo, chi))

    # === Thu thập style cho header cells ===
    header_cell_xf_indices = []
    for row_idx in range(header_row_count):
        row_xf = []
        for col_idx in range(total_cols):
            cell = source_sheet.cell(row_idx, col_idx)
            row_xf.append(cell.xf_index)
        header_cell_xf_indices.append(row_xf)

    # === Thu thập style cho data cells (dùng style từ dòng data đầu tiên) ===
    data_cell_xf_indices = []
    if header_row_count < total_rows:
        for col_idx in range(total_cols):
            cell = source_sheet.cell(header_row_count, col_idx)
            data_cell_xf_indices.append(cell.xf_index)

    # === Hàm tạo style xlwt từ xf_index gốc ===
    def make_xlwt_style(xf_index):
        """Tạo xlwt.XFStyle từ thông tin format gốc."""
        style = xlwt.XFStyle()

        try:
            xf = source_wb.xf_list[xf_index]
        except (IndexError, AttributeError):
            return style

        # Font
        try:
            src_font = source_wb.font_list[xf.font_index]
            font = xlwt.Font()
            font.name = src_font.name
            font.bold = src_font.bold
            font.italic = src_font.italic
            font.height = src_font.height
            font.underline = src_font.underline_type != 0
            font.colour_index = src_font.colour_index
            font.struck_out = src_font.struck_out
            style.font = font
        except (IndexError, AttributeError):
            pass

        # Alignment
        try:
            align = xlwt.Alignment()
            align.horz = xf.alignment.hor_align
            align.vert = xf.alignment.vert_align
            align.wrap = xf.alignment.text_wrap
            align.rota = xf.alignment.rotation
            style.alignment = align
        except AttributeError:
            pass

        # Border
        try:
            borders = xlwt.Borders()
            brd = xf.border
            borders.left = brd.left_line_style
            borders.right = brd.right_line_style
            borders.top = brd.top_line_style
            borders.bottom = brd.bottom_line_style
            borders.left_colour = brd.left_colour_index
            borders.right_colour = brd.right_colour_index
            borders.top_colour = brd.top_colour_index
            borders.bottom_colour = brd.bottom_colour_index
            style.borders = borders
        except AttributeError:
            pass

        # Background / Pattern
        try:
            pattern = xlwt.Pattern()
            bg = xf.background
            pattern.pattern = bg.fill_pattern
            pattern.pattern_fore_colour = bg.pattern_colour_index
            pattern.pattern_back_colour = bg.background_colour_index
            style.pattern = pattern
        except AttributeError:
            pass

        # Number format
        try:
            fmt_key = xf.format_key
            fmt_str = source_wb.format_map.get(fmt_key)
            if fmt_str:
                style.num_format_str = fmt_str.format_str
        except (AttributeError, KeyError):
            pass

        return style

    # === Cache styles ===
    style_cache = {}
    def get_style(xf_index):
        if xf_index not in style_cache:
            style_cache[xf_index] = make_xlwt_style(xf_index)
        return style_cache[xf_index]

    # === Pre-cache header styles ===
    header_styles = []
    for row_idx in range(header_row_count):
        row_styles = []
        for col_idx in range(total_cols):
            xf_idx = header_cell_xf_indices[row_idx][col_idx]
            row_styles.append(get_style(xf_idx))
        header_styles.append(row_styles)

    # === Pre-read header values ===
    header_values = []
    for row_idx in range(header_row_count):
        row_vals = []
        for col_idx in range(total_cols):
            val = source_sheet.cell_value(row_idx, col_idx)
            row_vals.append(val)
        header_values.append(row_vals)

    # === Pre-cache data styles (per column, from first data row) ===
    data_styles_per_col = []
    for col_idx in range(total_cols):
        if data_cell_xf_indices:
            data_styles_per_col.append(get_style(data_cell_xf_indices[col_idx]))
        else:
            data_styles_per_col.append(xlwt.XFStyle())

    created_files = []

    # === Tạo từng file ===
    for file_idx in range(num_files):
        data_start = header_row_count + file_idx * max_rows
        data_end = min(data_start + max_rows, total_rows)
        chunk_size = data_end - data_start

        output_filename = f"{source_name}_part{file_idx + 1}.xls"
        output_path = os.path.join(output_dir, output_filename)

        if progress_callback:
            pct = int((file_idx + 1) / num_files * 90) + 5
            progress_callback(pct, 100,
                f"Đang tạo file {file_idx + 1}/{num_files}: {output_filename} "
                f"({chunk_size:,} dòng dữ liệu)")

        # Tạo workbook mới
        out_wb = xlwt.Workbook(encoding='utf-8')
        out_sheet = out_wb.add_sheet(source_sheet.name or 'Sheet1')

        # === Ghi header rows (giữ nguyên format) ===
        for row_idx in range(header_row_count):
            for col_idx in range(total_cols):
                val = header_values[row_idx][col_idx]
                style = header_styles[row_idx][col_idx]
                out_sheet.write(row_idx, col_idx, val, style)

        # === Áp dụng merged cells cho header ===
        for (rlo, rhi, clo, chi) in header_merges:
            try:
                out_sheet.write_merge(
                    rlo, rhi - 1, clo, chi - 1,
                    header_values[rlo][clo],
                    header_styles[rlo][clo]
                )
            except Exception:
                pass

        # === Ghi data rows (giữ format theo cột) ===
        out_row = header_row_count
        for src_row in range(data_start, data_end):
            # Cố gắng lấy style riêng cho từng cell, fallback sang style theo cột
            for col_idx in range(total_cols):
                cell = source_sheet.cell(src_row, col_idx)
                val = cell.value

                # Xử lý date
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, source_wb.datemode)
                        val = val.strftime('%d/%m/%Y %H:%M:%S')
                    except Exception:
                        pass

                # Lấy style: ưu tiên style riêng của cell, fallback style cột
                try:
                    cell_xf = cell.xf_index
                    style = get_style(cell_xf)
                except Exception:
                    style = data_styles_per_col[col_idx]

                out_sheet.write(out_row, col_idx, val, style)

            # Cập nhật STT (cột 0) cho đúng thứ tự
            stt = (src_row - header_row_count) + 1
            try:
                out_sheet.write(out_row, 0, stt, get_style(
                    source_sheet.cell(src_row, 0).xf_index))
            except Exception:
                pass

            out_row += 1

        # === Đặt column widths ===
        for col_idx, width in col_widths.items():
            out_sheet.col(col_idx).width = width

        # === Đặt row heights cho header ===
        for row_idx, height in header_row_heights.items():
            out_sheet.row(row_idx).height_mismatch = True
            out_sheet.row(row_idx).height = height

        # === Freeze panes (đóng băng header) ===
        out_sheet.set_panes_frozen(True)
        out_sheet.set_horz_split_pos(header_row_count)

        out_wb.save(output_path)
        created_files.append(output_path)

    if progress_callback:
        progress_callback(100, 100,
            f"✅ Hoàn tất! Đã chia {data_row_count:,} dòng thành {num_files} file "
            f"trong thư mục:\n{output_dir}")

    return created_files


def split_excel(filepath, max_rows=MAX_ROWS_PER_FILE, progress_callback=None):
    """
    Wrapper chính - phân loại file và gọi hàm xử lý phù hợp.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        return split_xls_preserve_format(filepath, max_rows, progress_callback)
    elif ext in ('.xlsx', '.xlsm'):
        return _split_xlsx(filepath, max_rows, progress_callback)
    else:
        raise ValueError(f"Định dạng không hỗ trợ: {ext}\nChỉ hỗ trợ: .xls, .xlsx")


def _split_xlsx(filepath, max_rows=MAX_ROWS_PER_FILE, progress_callback=None):
    """Xử lý file .xlsx - đọc và ghi lại dạng .xls giữ data."""
    from openpyxl import load_workbook
    import xlwt

    if progress_callback:
        progress_callback(0, 100, "Đang đọc file .xlsx...")

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("File Excel trống!")

    # Tìm header: rows trước khi col 0 có số
    header_row_count = 0
    for i, row in enumerate(all_rows):
        if row[0] is not None and isinstance(row[0], (int, float)) and row[0] >= 1:
            header_row_count = i
            break
    if header_row_count == 0:
        header_row_count = 1

    headers = all_rows[:header_row_count]
    data_rows = all_rows[header_row_count:]
    total_cols = len(all_rows[0])

    if len(data_rows) <= max_rows:
        if progress_callback:
            progress_callback(100, 100, f"File chỉ có {len(data_rows):,} dòng, không cần chia!")
        return []

    num_files = math.ceil(len(data_rows) / max_rows)
    source_dir = os.path.dirname(os.path.abspath(filepath))
    source_name = os.path.splitext(os.path.basename(filepath))[0]
    output_dir = os.path.join(source_dir, f"{source_name}_split")
    os.makedirs(output_dir, exist_ok=True)

    # Basic styles
    header_style = xlwt.XFStyle()
    hf = xlwt.Font()
    hf.bold = True
    hf.name = 'Times New Roman'
    hf.height = 240
    header_style.font = hf
    hb = xlwt.Borders()
    hb.left = hb.right = hb.top = hb.bottom = xlwt.Borders.THIN
    header_style.borders = hb

    data_style = xlwt.XFStyle()
    df = xlwt.Font()
    df.name = 'Times New Roman'
    df.height = 200
    data_style.font = df
    data_style.borders = hb

    created_files = []
    for file_idx in range(num_files):
        start = file_idx * max_rows
        end = min(start + max_rows, len(data_rows))
        chunk = data_rows[start:end]

        output_filename = f"{source_name}_part{file_idx + 1}.xls"
        output_path = os.path.join(output_dir, output_filename)

        if progress_callback:
            pct = int((file_idx + 1) / num_files * 100)
            progress_callback(pct, 100,
                f"Đang tạo file {file_idx + 1}/{num_files}: {output_filename}")

        out_wb = xlwt.Workbook(encoding='utf-8')
        out_sheet = out_wb.add_sheet('Sheet1')

        # Write headers
        for r, row in enumerate(headers):
            for c in range(total_cols):
                val = row[c] if c < len(row) else ''
                out_sheet.write(r, c, val if val else '', header_style)

        # Write data
        for r, row in enumerate(chunk):
            for c in range(total_cols):
                val = row[c] if c < len(row) else ''
                out_sheet.write(header_row_count + r, c, val if val else '', data_style)

        out_sheet.set_panes_frozen(True)
        out_sheet.set_horz_split_pos(header_row_count)

        out_wb.save(output_path)
        created_files.append(output_path)

    if progress_callback:
        progress_callback(100, 100,
            f"✅ Hoàn tất! Đã chia thành {num_files} file trong: {output_dir}")

    return created_files


# ============================================================
# GUI
# ============================================================

class ExcelSplitterApp:
    """Giao diện đồ họa cho tool chia file Excel."""

    def __init__(self, root):
        self.root = root
        self.root.title("🗂️ Chia File Excel")
        self.root.geometry("700x550")
        self.root.resizable(True, True)

        self.is_running = False
        self.selected_file = tk.StringVar()
        self._build_ui()

    def _build_ui(self):
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(main_frame,
            text="TOOL CHIA FILE EXCEL", font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 5))

        desc_label = ttk.Label(main_frame,
            text=f"Chia file Excel thành các file .xls tối đa {MAX_ROWS_PER_FILE:,} dòng\n"
                 f"Giữ nguyên form mẫu gốc (font, border, màu nền, merged cells...)",
            font=('Arial', 10), justify=tk.CENTER)
        desc_label.pack(pady=(0, 15))

        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="File nguồn", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        file_entry = ttk.Entry(file_frame, textvariable=self.selected_file, state='readonly')
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        browse_btn = ttk.Button(file_frame, text="📂 Chọn file...", command=self._browse_file)
        browse_btn.pack(side=tk.RIGHT)

        # Config
        config_frame = ttk.LabelFrame(main_frame, text="Cấu hình", padding=10)
        config_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(config_frame, text="Số dòng dữ liệu tối đa mỗi file:").pack(side=tk.LEFT)
        self.max_rows_var = tk.StringVar(value=str(MAX_ROWS_PER_FILE))
        max_rows_entry = ttk.Entry(config_frame, textvariable=self.max_rows_var, width=10)
        max_rows_entry.pack(side=tk.LEFT, padx=(10, 0))

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.split_btn = ttk.Button(btn_frame, text="✂️ CHIA FILE", command=self._start_split)
        self.split_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.open_folder_btn = ttk.Button(btn_frame, text="📁 Mở thư mục kết quả",
            command=self._open_output_folder, state=tk.DISABLED)
        self.open_folder_btn.pack(side=tk.LEFT)

        # Progress
        self.progress = ttk.Progressbar(main_frame, mode='determinate', maximum=100)
        self.progress.pack(fill=tk.X, pady=(0, 10))

        # Log
        log_frame = ttk.LabelFrame(main_frame, text="Nhật ký", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD, state=tk.DISABLED,
                                font=('Consolas', 9))
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        self.output_dir = None

    def _browse_file(self):
        filepath = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[
                ("Excel files", "*.xls *.xlsx *.xlsm"),
                ("Excel 97-2003", "*.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        if filepath:
            self.selected_file.set(filepath)
            self._log(f"Đã chọn: {filepath}")

    def _log(self, message):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _update_progress(self, current, total, message):
        self.root.after(0, lambda: self._do_update(current, total, message))

    def _do_update(self, current, total, message):
        self.progress['value'] = current
        self._log(message)

    def _start_split(self):
        filepath = self.selected_file.get()
        if not filepath:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
            return

        try:
            max_rows = int(self.max_rows_var.get())
            if max_rows < 100:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Cảnh báo", "Số dòng tối đa phải là số nguyên >= 100!")
            return

        if self.is_running:
            return

        self.is_running = True
        self.split_btn.configure(state=tk.DISABLED)
        self.progress['value'] = 0

        thread = threading.Thread(target=self._do_split, args=(filepath, max_rows), daemon=True)
        thread.start()

    def _do_split(self, filepath, max_rows):
        try:
            created_files = split_excel(filepath, max_rows, self._update_progress)

            if created_files:
                self.output_dir = os.path.dirname(created_files[0])

                def on_done():
                    self.open_folder_btn.configure(state=tk.NORMAL)
                    self._log(f"\n{'='*50}")
                    self._log(f"✅ Tổng cộng: {len(created_files)} file đã được tạo")
                    self._log(f"📂 Thư mục: {self.output_dir}")
                    self._log(f"{'='*50}")
                    messagebox.showinfo("Thành công",
                        f"Đã chia thành {len(created_files)} file!\n\n"
                        f"Thư mục kết quả:\n{self.output_dir}")
                self.root.after(0, on_done)
            else:
                def on_no_split():
                    self._log("ℹ️ File không cần chia.")
                    messagebox.showinfo("Thông báo",
                        f"File có ít hơn {max_rows:,} dòng, không cần chia!")
                self.root.after(0, on_no_split)

        except Exception as e:
            def on_error():
                self._log(f"❌ Lỗi: {str(e)}")
                messagebox.showerror("Lỗi", f"Đã xảy ra lỗi:\n{str(e)}")
            self.root.after(0, on_error)

        finally:
            def reset():
                self.is_running = False
                self.split_btn.configure(state=tk.NORMAL)
            self.root.after(0, reset)

    def _open_output_folder(self):
        if self.output_dir and os.path.exists(self.output_dir):
            if sys.platform == 'win32':
                os.startfile(self.output_dir)
            elif sys.platform == 'darwin':
                os.system(f'open "{self.output_dir}"')
            else:
                os.system(f'xdg-open "{self.output_dir}"')


# ============================================================
# CLI
# ============================================================

def cli_mode(filepath):
    print(f"{'='*60}")
    print(f"  TOOL CHIA FILE EXCEL - Giữ nguyên form mẫu")
    print(f"  Tối đa {MAX_ROWS_PER_FILE:,} dòng mỗi file")
    print(f"{'='*60}")
    print(f"\n📄 File nguồn: {filepath}")

    def cli_progress(c, t, m):
        print(f"  [{c:3d}%] {m}")

    try:
        created_files = split_excel(filepath, MAX_ROWS_PER_FILE, cli_progress)
        if created_files:
            print(f"\n✅ Hoàn tất! Đã tạo {len(created_files)} file:")
            for f in created_files:
                print(f"   → {f}")
        else:
            print(f"\nℹ️  File không cần chia.")
    except Exception as e:
        print(f"\n❌ Lỗi: {e}")
        sys.exit(1)


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    if len(sys.argv) > 1:
        cli_mode(sys.argv[1])
    else:
        root = tk.Tk()
        app = ExcelSplitterApp(root)
        root.mainloop()
